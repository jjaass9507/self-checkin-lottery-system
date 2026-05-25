"""Small compatibility shim for the admin Excel import.

The project only used pandas for read_excel(...), df.columns and df.iterrows().
On Render Free, installing/importing real pandas + numpy is too memory-heavy.
This local module intentionally provides just the tiny subset the app needs,
backed by openpyxl read-only iteration.
"""

from openpyxl import load_workbook


class ExcelRow:
    def __init__(self, headers, values):
        self._data = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = values[index] if index < len(values) else None
            self._data[header] = value

    def __getitem__(self, key):
        return self._data[key]

    def get(self, key, default=None):
        return self._data.get(key, default)


class ExcelFrame:
    def __init__(self, workbook):
        self._workbook = workbook
        self._sheet = workbook.active
        self._rows = self._sheet.iter_rows(values_only=True)
        raw_headers = next(self._rows, ())
        self.columns = [str(value).strip() if value is not None else "" for value in raw_headers]

    def iterrows(self):
        try:
            for index, values in enumerate(self._rows):
                yield index, ExcelRow(self.columns, values)
        finally:
            self._workbook.close()


def read_excel(file):
    stream = getattr(file, "stream", file)
    if hasattr(stream, "seek"):
        stream.seek(0)
    workbook = load_workbook(filename=stream, read_only=True, data_only=True)
    return ExcelFrame(workbook)
