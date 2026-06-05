from flask import (
    Blueprint, render_template, request, flash, redirect, url_for, session, current_app, send_file
)
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from app import db
from app.models import CheckinList, DrawnTailNumber, Prize, CancellationLog, AppSetting
from datetime import datetime

bp = Blueprint('admin', __name__)

DASH_FIELD_KEYS = ['checkin_seq', 'site', 'dept', 'table', 'business_trip',
                   'participant_type', 'age_group', 'phone', 'meal', 'group', 'lottery_number',
                   'status', 'prize', 'checkin_time']
QUERY_FIELD_KEYS = ['employee_id', 'status', 'lottery_number', 'prize_info',
                    'table_number', 'meal_type', 'group_name', 'participant_type',
                    'age_group', 'phone']
DASH_FILTER_KEYS = ['status', 'site', 'dept', 'win', 'prize', 'type', 'meal', 'group']

ADMIN_PERMISSION_KEYS = [
    'import', 'reset_list', 'reset_checkin', 'reset_lottery',
    'manage_prizes', 'toggle_lottery', 'dash_config', 'query_config', 'test_checkin',
]

def _admin_role():
    return session.get('admin_role')

def _has_permission(feature):
    role = _admin_role()
    if role == 'super':
        return True
    if role == 'admin':
        s = AppSetting.query.get(f'admin_can_{feature}')
        return (s.value != 'false') if s else True
    return False

def _guard(feature):
    """Return a redirect response if current user lacks permission, else None."""
    if not _has_permission(feature):
        flash('您的帳號無此操作權限', 'danger')
        return redirect(url_for('admin.import_page'))
    return None

@bp.before_request
def require_login():
    if request.endpoint == 'admin.login':
        return
    if not session.get('admin_role'):
        return redirect(url_for('admin.login'))

@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pw = request.form.get('password', '')
        cfg = current_app.config
        if pw and pw == cfg.get('SUPER_ADMIN_PASSWORD'):
            role = 'super'
        elif pw and pw == cfg.get('ADMIN_PASSWORD'):
            role = 'admin'
        elif pw and cfg.get('VIEWER_PASSWORD') and pw == cfg.get('VIEWER_PASSWORD'):
            role = 'viewer'
        else:
            flash('密碼錯誤', 'danger')
            return render_template('admin/login.html')
        session['admin_role'] = role
        session['is_admin'] = True
        session.permanent = True
        flash('登入成功', 'success')
        if role == 'viewer':
            return redirect(url_for('admin.logs_page'))
        return redirect(url_for('admin.import_page'))
    return render_template('admin/login.html')

@bp.route('/logout')
def logout():
    session.pop('admin_role', None)
    session.pop('is_admin', None)
    flash('已登出', 'info')
    return redirect(url_for('admin.login'))

@bp.route('/')
def index():
    return redirect(url_for('admin.import_page'))

@bp.route('/import', methods=['GET', 'POST'])
def import_page():
    if _admin_role() == 'viewer':
        return redirect(url_for('admin.logs_page'))
    if request.method == 'POST':
        denied = _guard('import')
        if denied:
            return denied
        if 'file' not in request.files:
            flash('沒有檔案被上傳', 'danger')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('未選擇檔案', 'warning')
            return redirect(request.url)
        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file)
                if 'name' not in df.columns or 'employee_id' not in df.columns:
                    flash("Excel 檔案中缺少 'name' 或 'employee_id' 欄位", 'danger')
                    return redirect(request.url)
                imported_count = 0
                for _, row in df.iterrows():
                    name = str(row['name'])
                    employee_id = str(row['employee_id'])
                    lottery_number = str(row.get('lottery_number', ''))
                    site = str(row.get('site', ''))
                    dept_code = str(row.get('dept_code', ''))
                    table_val = str(row.get('table_number', ''))
                    table_number = table_val if table_val and table_val.lower() != 'nan' else None
                    biz_val = str(row.get('is_business_trip', '')).strip().upper()
                    is_business_trip = biz_val in ['1', 'Y', 'YES', 'TRUE', '是']
                    pt_raw = str(row.get('participant_type', '')).strip().lower()
                    if pt_raw in ['dependent', '眷屬', 'dep']:
                        participant_type = 'dependent'
                    elif pt_raw in ['employee', '員工', 'emp']:
                        participant_type = 'employee'
                    elif pt_raw in ['vendor_contact', '外部廠商主要窗口', '主要窗口']:
                        participant_type = 'vendor_contact'
                    elif pt_raw in ['vendor', '外部廠商', '廠商']:
                        participant_type = 'vendor'
                    else:
                        participant_type = None
                    linked_raw = str(row.get('linked_employee_id', '')).strip()
                    linked_employee_id = linked_raw if linked_raw and linked_raw.lower() != 'nan' else None
                    meal_raw = str(row.get('meal_type', '')).strip()
                    meal_type = meal_raw if meal_raw and meal_raw.lower() != 'nan' else None
                    group_raw = str(row.get('group_name', '')).strip()
                    group_name = group_raw if group_raw and group_raw.lower() != 'nan' else None
                    age_raw = str(row.get('age_group', '')).strip()
                    age_group = age_raw if age_raw and age_raw.lower() != 'nan' else None
                    phone_raw = str(row.get('phone', '')).strip()
                    phone = phone_raw if phone_raw and phone_raw.lower() != 'nan' else None
                    if not employee_id:
                        continue
                    exists = CheckinList.query.filter_by(employee_id=employee_id).first()
                    if not exists:
                        db.session.add(CheckinList(
                            name=name, employee_id=employee_id,
                            lottery_number=lottery_number if lottery_number else None,
                            site=site if site and site.lower() != 'nan' else None,
                            dept_code=dept_code if dept_code and dept_code.lower() != 'nan' else None,
                            table_number=table_number, is_business_trip=is_business_trip,
                            status='CheckedIn' if is_business_trip else 'Registered',
                            check_in_time=datetime.now() if is_business_trip else None,
                            participant_type=participant_type, linked_employee_id=linked_employee_id,
                            meal_type=meal_type, group_name=group_name, age_group=age_group, phone=phone,
                        ))
                        imported_count += 1
                db.session.commit()
                flash(f"成功！總共匯入了 {imported_count} 筆新的報到資料。", 'success')
            except Exception as e:
                db.session.rollback()
                flash(f"匯入時發生嚴重錯誤：{e}", 'danger')
            return redirect(url_for('admin.import_page'))

    query_fields = {}
    for key in QUERY_FIELD_KEYS:
        setting = AppSetting.query.get(f'query_show_{key}')
        query_fields[key] = (setting.value != 'false') if setting else True
    dash_fields = {}
    for key in DASH_FIELD_KEYS:
        setting = AppSetting.query.get(f'dash_show_{key}')
        dash_fields[key] = (setting.value != 'false') if setting else True
    filter_modes = {}
    for key in DASH_FILTER_KEYS:
        setting = AppSetting.query.get(f'dash_filter_mode_{key}')
        filter_modes[key] = setting.value if setting and setting.value in ('single', 'multiple') else 'single'
    # Admin permission settings (for super admin config panel)
    admin_perms = {}
    for key in ADMIN_PERMISSION_KEYS:
        s = AppSetting.query.get(f'admin_can_{key}')
        admin_perms[key] = (s.value != 'false') if s else True
    # Station enable/disable settings
    s_checkin = AppSetting.query.get('checkin_station_enabled')
    s_query = AppSetting.query.get('query_station_enabled')
    checkin_station_enabled = (s_checkin.value != 'false') if s_checkin else True
    query_station_enabled = (s_query.value != 'false') if s_query else True
    return render_template(
        'admin/import.html',
        query_fields=query_fields, dash_fields=dash_fields, filter_modes=filter_modes,
        admin_role=_admin_role(),
        admin_perms=admin_perms,
        checkin_station_enabled=checkin_station_enabled,
        query_station_enabled=query_station_enabled,
        has_permission=_has_permission,
    )

@bp.route('/dash_fields', methods=['POST'])
def toggle_dash_fields():
    denied = _guard('dash_config')
    if denied:
        return denied
    for key in DASH_FIELD_KEYS:
        new_value = 'true' if request.form.get(f'show_{key}') else 'false'
        setting = AppSetting.query.get(f'dash_show_{key}')
        if setting: setting.value = new_value
        else: db.session.add(AppSetting(key=f'dash_show_{key}', value=new_value))
    db.session.commit(); flash('報到名單欄位設定已更新', 'success')
    return redirect(url_for('admin.import_page'))

@bp.route('/dash_filter_modes', methods=['POST'])
def toggle_dash_filter_modes():
    denied = _guard('dash_config')
    if denied:
        return denied
    labels = {'status':'報到狀態','site':'站點','dept':'部門','win':'中獎狀態','prize':'特定獎項','type':'身分類別','meal':'餐點類型','group':'組別'}
    for key in DASH_FILTER_KEYS:
        value = request.form.get(f'filter_mode_{key}', 'single')
        value = 'multiple' if value == 'multiple' else 'single'
        setting_key = f'dash_filter_mode_{key}'
        setting = AppSetting.query.get(setting_key)
        if setting: setting.value = value
        else: db.session.add(AppSetting(key=setting_key, value=value))
    db.session.commit(); flash('報到儀表板篩選模式已更新', 'success')
    return redirect(url_for('admin.import_page'))

@bp.route('/query_fields', methods=['POST'])
def toggle_query_fields():
    denied = _guard('query_config')
    if denied:
        return denied
    for key in QUERY_FIELD_KEYS:
        new_value = 'true' if request.form.get(f'show_{key}') else 'false'
        setting = AppSetting.query.get(f'query_show_{key}')
        if setting: setting.value = new_value
        else: db.session.add(AppSetting(key=f'query_show_{key}', value=new_value))
    db.session.commit(); flash('查詢站欄位設定已更新', 'success')
    return redirect(url_for('admin.import_page'))

# --- 其餘管理功能 ---
@bp.route('/reset/list', methods=['POST'])
def reset_list():
    denied = _guard('reset_list')
    if denied:
        return denied
    try:
        deleted = db.session.query(CheckinList).delete(); db.session.commit(); flash(f"已清除所有名單 (共 {deleted} 筆)。", 'warning')
    except Exception as e:
        db.session.rollback(); flash(f"清除名單失敗：{e}", 'danger')
    return redirect(url_for('admin.import_page'))

@bp.route('/reset/checkin', methods=['POST'])
def reset_checkin():
    denied = _guard('reset_checkin')
    if denied:
        return denied
    try:
        updated = CheckinList.query.filter_by(status='CheckedIn').update({'status': 'Registered','check_in_time': None,'checkin_seq': None})
        db.session.commit(); flash(f"已重置所有報到狀態 (共 {updated} 人變回未報到)。", 'warning')
    except Exception as e:
        db.session.rollback(); flash(f"重置報到失敗：{e}", 'danger')
    return redirect(url_for('admin.import_page'))

@bp.route('/reset/lottery', methods=['POST'])
def reset_lottery():
    denied = _guard('reset_lottery')
    if denied:
        return denied
    try:
        deleted_log = db.session.query(DrawnTailNumber).delete()
        updated_rows = db.session.query(CheckinList).update({CheckinList.has_won: False, CheckinList.prize_claimed: False, CheckinList.has_won_public: False, CheckinList.public_prize_claimed: False})
        db.session.commit(); flash(f"已重置所有中獎紀錄 (清除 {deleted_log} 筆開獎，更新 {updated_rows} 人狀態)。", 'warning')
    except Exception as e:
        db.session.rollback(); flash(f"重置中獎失敗：{e}", 'danger')
    return redirect(url_for('admin.import_page'))

@bp.route('/reset/prize', methods=['POST'])
def reset_prize():
    prize_name = request.form.get('prize_name'); prize_type = request.form.get('prize_type')
    if not prize_name:
        flash('未指定獎項名稱', 'danger'); return redirect(url_for('admin.manage_prizes'))
    try:
        draws = DrawnTailNumber.query.filter_by(prize_name=prize_name, prize_type=prize_type).all()
        if not draws:
            flash(f"找不到「{prize_name}」的開獎紀錄，可能尚未抽出。", 'info'); return redirect(url_for('admin.manage_prizes'))
        draw_count = len(draws); updated_users = 0
        for draw in draws:
            suffix = str(draw.tail_number)
            if prize_type == 'tail' and not draw.is_addon:
                for p in CheckinList.query.filter(CheckinList.has_won == True, CheckinList.lottery_number.endswith(suffix)).all():
                    p.has_won = False; p.prize_claimed = False; updated_users += 1
            elif prize_type == 'public':
                target = str(draw.tail_number).zfill(3)
                for p in CheckinList.query.filter_by(has_won_public=True).all():
                    if (str(p.lottery_number).zfill(3) if p.lottery_number else '000') == target:
                        p.has_won_public = False; p.public_prize_claimed = False; updated_users += 1
            db.session.delete(draw)
        db.session.commit(); flash(f"已重置「{prize_name}」 (清除 {draw_count} 筆開獎紀錄，還原 {updated_users} 人為未中獎)。", 'success')
    except Exception as e:
        db.session.rollback(); flash(f"重置失敗：{e}", 'danger')
    return redirect(url_for('admin.manage_prizes'))

@bp.route('/test_checkin_all', methods=['POST'])
def test_checkin_all():
    denied = _guard('test_checkin')
    if denied:
        return denied
    try:
        pending_users = CheckinList.query.filter(CheckinList.status != 'CheckedIn').all(); count = len(pending_users)
        if count == 0:
            flash("目前所有人皆已報到。", 'info'); return redirect(url_for('admin.import_page'))
        current_time = datetime.now()
        for person in pending_users:
            person.status = 'CheckedIn'; person.check_in_time = current_time
        db.session.commit(); flash(f"【測試】已將剩餘 {count} 人標記為已報到。", 'success')
    except Exception as e:
        db.session.rollback(); flash(f"失敗：{e}", 'danger')
    return redirect(url_for('admin.import_page'))

@bp.route('/download_template')
def download_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '報到名單範本'
    headers = ['name','employee_id','lottery_number','site','dept_code','table_number','is_business_trip','participant_type','linked_employee_id','meal_type','group_name','age_group','phone']; ws.append(headers)
    ws.append(['王小明','A001','101','Taipei','IT','5','','employee','','A餐:便當','A組','大人','0912345678'])
    ws.append(['李小花','B002','202','','','','','dependent','A001','B餐:素食','A組','小孩',''])
    ws.append(['陳大雄','C003','303','Hsinchu','RD','8','Y','vendor','','C餐:滷味+綠豆冰沙','B組','大人',''])
    header_fill = PatternFill(start_color='0369a1', end_color='0369a1', fill_type='solid'); header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]: cell.fill = header_fill; cell.font = header_font
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col); ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='checkin_template.xlsx')

@bp.route('/toggle_lottery', methods=['POST'])
def toggle_lottery():
    denied = _guard('toggle_lottery')
    if denied:
        return denied
    new_value = request.form.get('lottery_enabled', 'true'); setting = AppSetting.query.get('lottery_enabled')
    if setting: setting.value = new_value
    else: db.session.add(AppSetting(key='lottery_enabled', value=new_value))
    db.session.commit(); flash(f"抽獎功能已{'啟用' if new_value == 'true' else '停用'}", 'success')
    return redirect(url_for('admin.import_page'))

@bp.route('/logs')
def logs_page():
    logs = db.session.query(CancellationLog, CheckinList).join(CheckinList, CancellationLog.checkin_list_id == CheckinList.id).order_by(CancellationLog.timestamp.desc()).all()
    return render_template('admin/logs.html', logs=logs)

@bp.route('/prizes', methods=['GET', 'POST'])
def manage_prizes():
    if request.method == 'POST':
        denied = _guard('manage_prizes')
        if denied:
            return denied
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('name'); prize_type = request.form.get('prize_type'); quantity = int(request.form.get('quantity', 1)); display_order = int(request.form.get('display_order', 0))
            if name: db.session.add(Prize(name=name, prize_type=prize_type, quantity=quantity, display_order=display_order)); flash('獎項已新增', 'success')
        elif action == 'edit':
            prize = Prize.query.get(request.form.get('prize_id'))
            if prize:
                prize.name = request.form.get('name'); prize.prize_type = request.form.get('prize_type'); prize.quantity = int(request.form.get('quantity', 1)); prize.display_order = int(request.form.get('display_order', 0)); flash('獎項已更新', 'success')
        elif action == 'delete':
            prize = Prize.query.get(request.form.get('prize_id'))
            if prize: db.session.delete(prize); flash('獎項已刪除', 'warning')
        db.session.commit(); return redirect(url_for('admin.manage_prizes'))
    prizes = Prize.query.order_by(Prize.prize_type, Prize.display_order).all()
    return render_template('admin/prizes.html', prizes=prizes)

@bp.route('/admin_permissions', methods=['POST'])
def save_admin_permissions():
    if _admin_role() != 'super':
        flash('僅超級管理員可修改此設定', 'danger')
        return redirect(url_for('admin.import_page'))
    for key in ADMIN_PERMISSION_KEYS:
        new_value = 'true' if request.form.get(f'admin_can_{key}') else 'false'
        setting = AppSetting.query.get(f'admin_can_{key}')
        if setting:
            setting.value = new_value
        else:
            db.session.add(AppSetting(key=f'admin_can_{key}', value=new_value))
    db.session.commit()
    flash('管理員權限設定已更新', 'success')
    return redirect(url_for('admin.import_page'))

@bp.route('/toggle_stations', methods=['POST'])
def toggle_stations():
    if _admin_role() != 'super':
        flash('僅超級管理員可修改此設定', 'danger')
        return redirect(url_for('admin.import_page'))
    for key in ('checkin_station_enabled', 'query_station_enabled'):
        new_value = request.form.get(key, 'false')
        new_value = 'true' if new_value == 'true' else 'false'
        setting = AppSetting.query.get(key)
        if setting:
            setting.value = new_value
        else:
            db.session.add(AppSetting(key=key, value=new_value))
    db.session.commit()
    flash('前台站點設定已更新', 'success')
    return redirect(url_for('admin.import_page'))
