import os
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile

import openpyxl
from flask import flash, redirect, render_template, request, send_file, url_for
from openpyxl.styles import Font, PatternFill

from app import db
from app.admin.routes import (
    bp, _admin_role, _has_permission, _guard,
    ADMIN_PERMISSION_KEYS, DASH_FILTER_KEYS,
)
from app.admin.xlsx_stream import iter_xlsx_rows
from app.models import AppSetting, CancellationLog, CheckinList

IMPORT_BATCH_SIZE = 200

PARTICIPANT_TYPE_ALIASES = {
    'employee': 'employee', 'emp': 'employee', '員工': 'employee', '同仁': 'employee',
    'dependent': 'dependent', 'dep': 'dependent', '眷屬': 'dependent', '家屬': 'dependent',
    'vendor_contact': 'vendor_contact', 'external_vendor_contact': 'vendor_contact',
    'vendor_main_contact': 'vendor_contact', 'main_vendor_contact': 'vendor_contact',
    '主要窗口': 'vendor_contact', '主窗口': 'vendor_contact', '廠商主要窗口': 'vendor_contact',
    '廠商主窗口': 'vendor_contact', '外部廠商主要窗口': 'vendor_contact', '外部廠商主窗口': 'vendor_contact',
    '外部廠商主要聯絡人': 'vendor_contact', '外部廠商主聯絡人': 'vendor_contact',
    'vendor': 'vendor', 'external_vendor': 'vendor', '外部廠商': 'vendor', '廠商': 'vendor',
}

QUERY_FIELD_KEYS = [
    'employee_id', 'status', 'lottery_number', 'prize_info', 'table_number',
    'meal_type', 'group_name', 'participant_type', 'age_group', 'phone'
]
DASH_FIELD_KEYS = [
    'checkin_seq', 'site', 'dept', 'table', 'business_trip', 'participant_type',
    'age_group', 'phone', 'meal', 'group', 'lottery_number', 'status', 'prize', 'checkin_time'
]


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    if text.endswith('.0') and text[:-2].isdigit():
        text = text[:-2]
    if not text or text.lower() == 'nan':
        return None
    return text


def _get(row, headers, key):
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    return _clean(row[idx])


def _parse_business_trip(value):
    value = _clean(value)
    return bool(value and value.upper() in ['1', 'Y', 'YES', 'TRUE', '是'])


def _canonical_key(value):
    value = (_clean(value) or '').strip()
    return value.lower().replace('-', '_').replace(' ', '_')


def _parse_participant_type(value):
    raw = (_clean(value) or '').strip()
    compact = ''.join(raw.split())
    normalized = _canonical_key(raw)
    normalized_compact = _canonical_key(compact)
    return (
        PARTICIPANT_TYPE_ALIASES.get(raw)
        or PARTICIPANT_TYPE_ALIASES.get(raw.lower())
        or PARTICIPANT_TYPE_ALIASES.get(compact)
        or PARTICIPANT_TYPE_ALIASES.get(compact.lower())
        or PARTICIPANT_TYPE_ALIASES.get(normalized)
        or PARTICIPANT_TYPE_ALIASES.get(normalized_compact)
    )


def _next_dependent_employee_id(original_employee_id, dependent_serials):
    current = dependent_serials.get(original_employee_id, 0) + 1
    dependent_serials[original_employee_id] = current
    return f'{original_employee_id}_{current}'


def _build_row_data(row, headers, dependent_serials):
    name = _get(row, headers, 'name')
    raw_employee_id = _get(row, headers, 'employee_id')
    if not name or not raw_employee_id:
        return None

    participant_type = _parse_participant_type(_get(row, headers, 'participant_type'))
    employee_id = raw_employee_id
    linked_employee_id = _get(row, headers, 'linked_employee_id')

    if participant_type == 'dependent':
        linked_employee_id = raw_employee_id
        employee_id = _next_dependent_employee_id(raw_employee_id, dependent_serials)

    is_business_trip = _parse_business_trip(_get(row, headers, 'is_business_trip'))

    return {
        'name': name,
        'employee_id': employee_id,
        'lottery_number': _get(row, headers, 'lottery_number'),
        'site': _get(row, headers, 'site'),
        'dept_code': _get(row, headers, 'dept_code'),
        'table_number': _get(row, headers, 'table_number'),
        'is_business_trip': is_business_trip,
        'status': 'CheckedIn' if is_business_trip else 'Registered',
        'check_in_time': datetime.now() if is_business_trip else None,
        'participant_type': participant_type,
        'linked_employee_id': linked_employee_id,
        'meal_type': _clean(_get(row, headers, 'meal_type')),
        'group_name': _get(row, headers, 'group_name'),
        'age_group': _get(row, headers, 'age_group'),
        'phone': _get(row, headers, 'phone'),
    }


def _save_batch(rows):
    if not rows:
        return 0, 0
    employee_ids = [row['employee_id'] for row in rows]
    existing_ids = {
        employee_id for (employee_id,) in db.session.query(CheckinList.employee_id)
        .filter(CheckinList.employee_id.in_(employee_ids)).all()
    }
    new_items = []
    skipped = 0
    for row_data in rows:
        employee_id = row_data['employee_id']
        if employee_id in existing_ids:
            skipped += 1
            continue
        new_items.append(CheckinList(**row_data))
        existing_ids.add(employee_id)
    if new_items:
        db.session.add_all(new_items)
        db.session.commit()
        db.session.expunge_all()
    return len(new_items), skipped


def _render_import_page():
    query_fields = {}
    for key in QUERY_FIELD_KEYS:
        setting = AppSetting.query.get(f'query_show_{key}')
        query_fields[key] = (setting.value != 'false') if setting else True

    dash_fields = {}
    for key in DASH_FIELD_KEYS:
        setting = AppSetting.query.get(f'dash_show_{key}')
        dash_fields[key] = (setting.value != 'false') if setting else True

    filter_modes = {}
    for key in DASH_FILTER_KEYS:
        setting = AppSetting.query.get(f'dash_filter_mode_{key}')
        filter_modes[key] = setting.value if setting and setting.value in ('single', 'multiple') else 'single'

    admin_perms = {}
    for key in ADMIN_PERMISSION_KEYS:
        s = AppSetting.query.get(f'admin_can_{key}')
        admin_perms[key] = (s.value != 'false') if s else True

    s_checkin = AppSetting.query.get('checkin_station_enabled')
    s_query = AppSetting.query.get('query_station_enabled')
    checkin_station_enabled = (s_checkin.value != 'false') if s_checkin else True
    query_station_enabled = (s_query.value != 'false') if s_query else True

    return render_template(
        'admin/import.html',
        query_fields=query_fields, dash_fields=dash_fields, filter_modes=filter_modes,
        admin_role=_admin_role(),
        admin_perms=admin_perms,
        checkin_station_enabled=checkin_station_enabled,
        query_station_enabled=query_station_enabled,
        has_permission=_has_permission,
    )


def optimized_import_page():
    if _admin_role() == 'viewer':
        return redirect(url_for('admin.logs_page'))
    if request.method != 'POST':
        return _render_import_page()

    denied = _guard('import')
    if denied:
        return denied

    if 'file' not in request.files:
        flash('沒有檔案被上傳', 'danger')
        return redirect(request.url)
    file = request.files['file']
    if file.filename == '':
        flash('未選擇檔案', 'warning')
        return redirect(request.url)
    if not file.filename.lower().endswith('.xlsx'):
        flash('請上傳 .xlsx 格式的 Excel 檔案', 'warning')
        return redirect(request.url)

    tmp_path = None
    imported_count = 0
    skipped_existing_count = 0
    skipped_invalid_count = 0

    try:
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp)
            tmp_path = tmp.name

        row_iter = iter_xlsx_rows(tmp_path)
        raw_headers = next(row_iter, None)
        if not raw_headers:
            flash('Excel 檔案是空的', 'danger')
            return redirect(request.url)

        headers = {str(header).strip(): index for index, header in enumerate(raw_headers) if header is not None and str(header).strip()}
        if 'name' not in headers or 'employee_id' not in headers:
            flash("Excel 檔案中缺少 'name' 或 'employee_id' 欄位", 'danger')
            return redirect(request.url)

        batch = []
        dependent_serials = {}
        for row in row_iter:
            row_data = _build_row_data(row, headers, dependent_serials)
            if not row_data:
                skipped_invalid_count += 1
                continue
            batch.append(row_data)
            if len(batch) >= IMPORT_BATCH_SIZE:
                inserted, skipped = _save_batch(batch)
                imported_count += inserted
                skipped_existing_count += skipped
                batch.clear()

        if batch:
            inserted, skipped = _save_batch(batch)
            imported_count += inserted
            skipped_existing_count += skipped

        message = f"成功！總共匯入了 {imported_count} 筆新的報到資料。"
        if skipped_existing_count:
            message += f" 已跳過 {skipped_existing_count} 筆既有或重複工號。"
        if skipped_invalid_count:
            message += f" 已跳過 {skipped_invalid_count} 筆缺少姓名或工號的資料。"
        flash(message, 'success')
    except Exception as exc:
        db.session.rollback()
        if imported_count:
            flash(f"匯入時發生錯誤：{exc}。已成功寫入 {imported_count} 筆，請檢查資料後再補匯。", 'danger')
        else:
            flash(f"匯入時發生嚴重錯誤：{exc}", 'danger')
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
    return redirect(url_for('admin.import_page'))


def optimized_reset_list():
    denied = _guard('reset_list')
    if denied:
        return denied
    try:
        deleted_logs = db.session.query(CancellationLog).delete(synchronize_session=False)
        deleted_list = db.session.query(CheckinList).delete(synchronize_session=False)
        db.session.commit()
        flash(f"已清除所有名單 (共 {deleted_list} 筆)，並清除取消紀錄 {deleted_logs} 筆。", 'warning')
    except Exception as exc:
        db.session.rollback()
        flash(f"清除名單失敗：{exc}", 'danger')
    return redirect(url_for('admin.import_page'))


def optimized_toggle_dash_fields():
    denied = _guard('dash_config')
    if denied:
        return denied
    for key in DASH_FIELD_KEYS:
        new_value = 'true' if request.form.get(f'show_{key}') else 'false'
        setting = AppSetting.query.get(f'dash_show_{key}')
        if setting:
            setting.value = new_value
        else:
            db.session.add(AppSetting(key=f'dash_show_{key}', value=new_value))
    db.session.commit()
    flash('報到名單欄位設定已更新', 'success')
    return redirect(url_for('admin.import_page'))


def optimized_toggle_query_fields():
    denied = _guard('query_config')
    if denied:
        return denied
    for key in QUERY_FIELD_KEYS:
        new_value = 'true' if request.form.get(f'show_{key}') else 'false'
        setting = AppSetting.query.get(f'query_show_{key}')
        if setting:
            setting.value = new_value
        else:
            db.session.add(AppSetting(key=f'query_show_{key}', value=new_value))
    db.session.commit()
    flash('查詢站欄位設定已更新', 'success')
    return redirect(url_for('admin.import_page'))


def optimized_download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '報到名單範本'
    headers = [
        'name', 'employee_id', 'lottery_number', 'site', 'dept_code', 'table_number',
        'is_business_trip', 'participant_type', 'linked_employee_id', 'meal_type',
        'group_name', 'age_group', 'phone'
    ]
    ws.append(headers)
    ws.append(['王小明', 'A001', '101', 'Taipei', 'IT部門', '5', '', 'employee', '', 'A', '第一組', '大人', '0912345678'])
    ws.append(['李小花', 'A001', '202', '', '', '', '', 'dependent', '', 'C餐:滷味+綠豆冰沙', '第一組', '小孩', ''])
    ws.append(['陳窗口', 'V001', '303', 'Hsinchu', '外包服務', '8', '', '外部廠商主要窗口', '', 'B', '廠商組', '大人', '0987654321'])
    ws.append(['林廠商', 'V002', '304', 'Hsinchu', '外包服務', '8', '', '外部廠商', '', 'B', '廠商組', '大人', ''])

    header_fill = PatternFill(start_color='0369a1', end_color='0369a1', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='checkin_template.xlsx')


bp.view_functions['import_page'] = optimized_import_page
bp.view_functions['reset_list'] = optimized_reset_list
bp.view_functions['toggle_dash_fields'] = optimized_toggle_dash_fields
bp.view_functions['toggle_query_fields'] = optimized_toggle_query_fields
bp.view_functions['download_template'] = optimized_download_template
